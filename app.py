from ctypes import wstring_at
from openpyxl import Workbook, load_workbook
import datetime

def main():
    # Create the workbook object from our data file
    wb = load_workbook(filename = '.\Data\OG Meter Data Summary.xlsx')

    # Save this to a new file so that we leave the OG data intact
    wb.save('.\Data\Updated Meter Data Summary.xlsx')

    # Reassign wb to the new updated meter data
    wb = load_workbook(filename = '.\Data\Updated Meter Data Summary.xlsx')

    # Iterate through each worksheet (ws) in the workbook (wb)
    for ws in wb:
        print(ws.title)

        # Skip over index sheet and HDD/CDD sheet
        if ws.title == "Index Sheet" or ws.title == "HDD&CDD":
            continue

        # Confirm that the B column is the DATE column. If not then skip this sheet.
        if ws['B1'].value == "DATE":

            # Iterate through all of the cells in column B
            for i, cell in enumerate(ws['B']):

                # Remove any rows that do not have a value in the DATE column
                if cell.value == None:
                    print("Row "+str(i)+" deleted.")
                    ws.delete_rows(i+1, 1)

            # Save any changes made to this column
            wb.save('.\Data\Updated Meter Data Summary.xlsx')
        else:
            print("Date column FAILED")
            continue

        # Confirm that the C column is the METER READING column. If not then skip this sheet.
        if ws["C1"].value == "METER READING":

            # Iterate through all of the cells in column C
            for i, cell in enumerate(ws['C']):

                # Remove any rows that do not have a value in the METER READING column
                if cell.value == None or cell.value == "Missing":
                    print("Row "+str(i)+" deleted.")
                    ws.delete_rows(i+1, 1)

            # Save any changes made to this column
            wb.save('.\Data\Updated Meter Data Summary.xlsx')

        else:
            print("Meter reading column FAILED")
            continue

        # Confirm that the D column is the CONSUMPTION column. If not then skip this sheet.
        if ws["D1"].value == "CONSUMPTION":

            # Iterate through all of the cells in column D
            for i, cell in enumerate(ws['C']):

                # Remove any rows that do not have a value in the CONSUMPTION column
                if cell.value == None or cell.value == "Missing":
                    print("Row "+str(i)+" deleted.")
                    ws.delete_rows(i+1, 1)

            # Save any changes made to this column
            wb.save('.\Data\Updated Meter Data Summary.xlsx')

        else:
            print("Consumption column FAILED")
            continue

        print('----------')

    # One last save
    wb.save('.\Data\Updated Meter Data Summary.xlsx')

    """ ws = wb['TN Elec']

    index = 0

    for i, cell in enumerate(ws['B']):
        print(cell.number_format)
        print(cell.value)
        index = i

    print('Number of rows: '+str(i)) """

if __name__ == "__main__":
    main()